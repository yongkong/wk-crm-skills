# -*- coding: utf-8 -*-
"""
CRM 新模块需求模板 Excel 填充脚本

使用方法：
1. 复制本脚本到工作目录
2. 修改 DATA 字典中的数据
3. 执行：python fill_xlsx.py

输入：空白模板 CRM新模块开发需求模板.xlsx
输出：填写后的 CRM新模块开发需求模板.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from pathlib import Path

# ============================================================
# 配置区 - 修改以下数据
# ============================================================

# 模板路径（相对于项目根目录）
TEMPLATE_PATH = Path(r"wk-crm-skills/wk-crm-new-module/references/CRM新模块开发需求模板.xlsx")
OUTPUT_PATH = Path(r"doc/{模块目录}/CRM新模块开发需求模板.xlsx")

# 绿色样式（标识已填写数据）
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(name="微软雅黑", size=10, color="006100")

# ============================================================
# 数据定义 - 按模块修改
# ============================================================

DATA = {
    # Sheet 1: 模块基础信息
    "base_info": [
        # (参数名, 参数值, 填写说明)
        ("模块中文名", "{模块中文名}", "中文显示名"),
        ("模块英文名(驼峰)", "{moduleCamelCase}", "用于生成类名/文件名/路由"),
        ("CrmEnum.type", "{type}", "已在 module-registry.md 注册"),
        ("菜单基址", "{menuBase}", "已在 module-registry.md 注册"),
        ("主键字段名", "{pkName}", "驼峰，= 模块英文名 + Id"),
        ("编号字段名", "{numFieldName}", "业务编号"),
        ("是否需要审批流", "{needExamine}", "L2+必填"),
        ("表单渲染模式", "{renderMode}", "A/B/C"),
        ("参考模块", "{refModule}", "最接近的已有模块"),
        ("需求文档路径", "{docPath}", "需求文档目录"),
    ],

    # Sheet 2: 主表业务字段（系统字段已在模板中预填）
    "main_fields": [
        # (字段名, 显示名, 数据库类型, 必填, 默认值, 说明)
        ("{field_name}", "{显示名}", "varchar(50)", "是", "NULL", "说明"),
    ],

    # Sheet 3: 表单字段
    "form_fields": [
        # (字段名, 显示名, formType, 必填, 选项/规则, 列表显示, 系统字段, 说明)
        ("fieldName", "显示名", "text", "是", "", "是", "否", "说明"),
    ],

    # Sheet 4: 子表字段（L3+ 填写）
    "sub_fields": [
        # (子表名, 子表中文名, 字段名, 字段显示名, 类型, 必填, 说明)
        ("sub_table", "子表中文名", "fieldName", "字段显示名", "text", "是", "说明"),
    ],

    # Sheet 5: 业务规则
    "rules": [
        # (规则类型, 规则描述, 触发时机, 说明)
        ("状态流转", "草稿→审批中→通过/驳回", "保存/提交/审批回调", "checkStatus 状态机"),
        ("自动填充", "选择合同后自动填充相关字段", "表单 change 事件", "减少手工输入"),
        ("计算规则", "计算公式描述", "表单 change 事件", "前端实时计算"),
        ("校验规则", "校验规则描述", "提交前校验", "防止重复操作"),
        ("ERP集成", "ERP同步描述", "审批回调", "L3+级别ERP同步"),
        ("编号规则", "格式：XX-{yyMMdd}-{4位序号}", "新建保存时", "serial_number类型字段"),
    ],

    # Sheet 6: 页面与权限
    "page_config": [
        # (配置项, 值, 说明)
        ("列表Tab", "全部,我负责的,下属的,我关注的", "标准四个 Tab"),
        ("列表默认字段", "编号,类型,状态,创建时间", "从表单字段中选择"),
        ("操作按钮", "新建,编辑,删除,转移", "标准CRUD"),
        ("是否需要详情页", "是", "复杂模块需要详情页"),
        ("额外权限按钮", "作废,撤回", "自定义按钮"),
        ("授权角色", "管理员(全部)", "管理员拥有全部权限"),
    ],
}


# ============================================================
# 工具函数
# ============================================================

def unmerge_data_rows(ws, start_row=3, max_row=50):
    """Unmerge any merged cells in data rows area."""
    to_unmerge = []
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= start_row and mc.min_row <= max_row:
            to_unmerge.append(str(mc))
    for mc_str in to_unmerge:
        ws.unmerge_cells(mc_str)


def clear_rows(ws, start_row, max_row=None):
    """Clear cell values in a row range, skipping merged cells."""
    if max_row is None:
        max_row = ws.max_row
    for row in ws.iter_rows(min_row=start_row, max_row=max_row, max_col=ws.max_column):
        for cell in row:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = None


def style_row(ws, row, fill=GREEN_FILL, font=GREEN_FONT):
    """Apply green style to filled data row."""
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font


# ============================================================
# 填充函数
# ============================================================

def fill_sheet1(wb, data):
    """填充 Sheet 1: 模块基础信息"""
    ws = wb["模块基础信息"]
    unmerge_data_rows(ws, start_row=3)
    clear_rows(ws, 3)

    for i, (name, value, note) in enumerate(data["base_info"], start=3):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=value)
        ws.cell(row=i, column=3, value=note)
        style_row(ws, i)

    print("✓ Sheet 1: 模块基础信息")


def fill_sheet2(wb, data):
    """填充 Sheet 2: 主表字段（业务字段从第 12 行开始）"""
    ws = wb["主表字段"]
    unmerge_data_rows(ws, start_row=12)
    clear_rows(ws, 12)

    for i, (fname, display, dtype, req, default, desc) in enumerate(data["main_fields"], start=12):
        ws.cell(row=i, column=1, value=fname)
        ws.cell(row=i, column=2, value=display)
        ws.cell(row=i, column=3, value=dtype)
        ws.cell(row=i, column=4, value=req)
        ws.cell(row=i, column=5, value=default)
        ws.cell(row=i, column=6, value="否")  # 是否系统字段
        ws.cell(row=i, column=7, value=desc)
        style_row(ws, i)

    print("✓ Sheet 2: 主表字段")


def fill_sheet3(wb, data):
    """填充 Sheet 3: 表单字段"""
    ws = wb["表单字段"]
    unmerge_data_rows(ws, start_row=3)
    clear_rows(ws, 3)

    for i, (fname, display, ftype, req, opts, list_show, sys_auto, desc) in enumerate(data["form_fields"], start=3):
        ws.cell(row=i, column=1, value=fname)
        ws.cell(row=i, column=2, value=display)
        ws.cell(row=i, column=3, value=ftype)
        ws.cell(row=i, column=4, value=req)
        ws.cell(row=i, column=5, value=opts)
        ws.cell(row=i, column=6, value=list_show)
        ws.cell(row=i, column=7, value=sys_auto)
        ws.cell(row=i, column=8, value=desc)
        style_row(ws, i)

    print("✓ Sheet 3: 表单字段")


def fill_sheet4(wb, data):
    """填充 Sheet 4: 子表设计"""
    ws = wb["子表设计"]
    unmerge_data_rows(ws, start_row=3)
    clear_rows(ws, 3)

    for i, (tname, tcn, fname, fcn, ftype, req, desc) in enumerate(data["sub_fields"], start=3):
        ws.cell(row=i, column=1, value=tname)
        ws.cell(row=i, column=2, value=tcn)
        ws.cell(row=i, column=3, value=fname)
        ws.cell(row=i, column=4, value=fcn)
        ws.cell(row=i, column=5, value=ftype)
        ws.cell(row=i, column=6, value=req)
        ws.cell(row=i, column=7, value=desc)
        style_row(ws, i)

    print("✓ Sheet 4: 子表设计")


def fill_sheet5(wb, data):
    """填充 Sheet 5: 业务规则"""
    ws = wb["业务规则"]
    unmerge_data_rows(ws, start_row=3)
    clear_rows(ws, 3)

    for i, (rtype, rdesc, trigger, note) in enumerate(data["rules"], start=3):
        ws.cell(row=i, column=1, value=rtype)
        ws.cell(row=i, column=2, value=rdesc)
        ws.cell(row=i, column=3, value=trigger)
        ws.cell(row=i, column=4, value=note)
        style_row(ws, i)

    print("✓ Sheet 5: 业务规则")


def fill_sheet6(wb, data):
    """填充 Sheet 6: 页面与权限"""
    ws = wb["页面与权限"]
    unmerge_data_rows(ws, start_row=3)
    clear_rows(ws, 3)

    for i, (item, value, note) in enumerate(data["page_config"], start=3):
        ws.cell(row=i, column=1, value=item)
        ws.cell(row=i, column=2, value=value)
        ws.cell(row=i, column=3, value=note)
        style_row(ws, i)

    print("✓ Sheet 6: 页面与权限")


# ============================================================
# 主函数
# ============================================================

def main():
    # 加载模板
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    print(f"Loaded template: {TEMPLATE_PATH}")

    # 填充各 Sheet
    fill_sheet1(wb, DATA)
    fill_sheet2(wb, DATA)
    fill_sheet3(wb, DATA)
    fill_sheet4(wb, DATA)
    fill_sheet5(wb, DATA)
    fill_sheet6(wb, DATA)

    # 保存
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"\n✅ Excel saved to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
